import csv
import io
import logging
import re
from datetime import datetime
from .models import PreviewData

logger = logging.getLogger(__name__)


def _read_csv_rows(file_path):
    """Read CSV file and return list of dicts."""
    try:
        if hasattr(file_path, 'read'):
            content = file_path.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8-sig')
            else:
                content = str(content)
            file_path.seek(0)
        else:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()
    except Exception:
        content = ''

    reader = csv.DictReader(io.StringIO(content))
    return [
        {k.strip(): v.strip() if v else None for k, v in row.items()}
        for row in reader
    ]


def _read_excel_rows(file_path):
    """Read Excel file using openpyxl and return list of dicts with column headers."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(cell).strip() if cell is not None else '' for cell in row]
        else:
            row_dict = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers):
                    key = headers[col_idx]
                    if cell_value is not None:
                        row_dict[key] = str(cell_value).strip()
                    else:
                        row_dict[key] = None
            if any(v is not None for v in row_dict.values()):
                rows.append(row_dict)

    wb.close()
    return rows


def _coerce_numeric(value):
    """Convert a value to float, return None if not possible."""
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


class DataProcessor:
    """Handles data processing and validation."""

    REQUIRED_FIELDS = [
        'Well',
        'Date',
        'BS&W (%)',
        'Net Oil (bopd)',
        'Form.GLR (scf/bbl)',
        'Prod Method',
        'Test Status',
        'Tubing Pressure (psi)',
        'Flow Line Pressure (psi)',
        'Well Choke Size',
    ]

    FIELD_ALIASES = {
        'Well': ['well', 'wellname', 'well_id', 'wellid', 'wellnameid', 'well name'],
        'Date': ['date', 'testdate', 'datetime', 'date_time', 'test_date'],
        'BS&W (%)': ['bsw', 'bs&w', 'watercut', 'wc', 'watercutpercent', 'watercut%'],
        'Net Oil (bopd)': ['netoil', 'oilrate', 'oil_bopd', 'oilratebopd', 'oilbopd'],
        'Form.GLR (scf/bbl)': ['glr', 'formglr', 'gasliquidratio', 'gasliquidratio_scf_bbl'],
        'Prod Method': ['prodmethod', 'productionmethod', 'production_method', 'method'],
        'Test Status': ['teststatus', 'status', 'test_status'],
        'Tubing Pressure (psi)': ['tubingpressure', 'thp', 'tp', 'tubingpressurepsi'],
        'Flow Line Pressure (psi)': ['flowlinepressure', 'flp', 'fp', 'flowlinepressurepsi'],
        'Well Choke Size': ['wellchokesize', 'chokesize', 'choke', 'wellchoke', 'choke size'],
    }

    @staticmethod
    def read_file(file_obj, file_format):
        """Read file and return list of dict rows."""
        if file_format == 'xlsx':
            return _read_excel_rows(file_obj)
        else:
            return _read_csv_rows(file_obj)

    @staticmethod
    def detect_columns(rows):
        """Detect available columns from a list of dict rows."""
        if not rows:
            return []
        return list(rows[0].keys())

    @staticmethod
    def normalize_name(value):
        if value is None:
            return ''
        return re.sub(r'[^a-z0-9]+', '', str(value).lower())

    @classmethod
    def auto_map_columns(cls, available_columns):
        """Suggest likely column matches for the required fields."""
        mapping = {}
        available_columns = [col for col in available_columns if col is not None]

        for field in cls.REQUIRED_FIELDS:
            best_match = None
            best_score = 0
            aliases = cls.FIELD_ALIASES.get(field, [field])
            normalized_aliases = [cls.normalize_name(alias) for alias in aliases]

            for column in available_columns:
                normalized_column = cls.normalize_name(column)
                if not normalized_column:
                    continue

                score = 0
                if normalized_column in normalized_aliases:
                    score = 100
                else:
                    for alias in normalized_aliases:
                        if alias and alias in normalized_column:
                            score = max(score, 90)
                        elif normalized_column.startswith(alias) or normalized_column.endswith(alias):
                            score = max(score, 80)
                        elif alias and len(alias) > 3 and alias in normalized_column:
                            score = max(score, 70)

                    if score == 0:
                        field_tokens = [token for token in re.split(r'[^a-z0-9]+', field.lower()) if token and token not in {'and', 'the'}]
                        column_tokens = [token for token in re.split(r'[^a-z0-9]+', normalized_column) if token]
                        overlap = len(set(field_tokens) & set(column_tokens))
                        if overlap and overlap >= max(1, len(field_tokens) // 2):
                            score = 60

                if score > best_score:
                    best_score = score
                    best_match = column

            if best_score >= 70:
                mapping[field] = best_match

        return mapping

    @staticmethod
    def validate_mapping(mapping):
        """Validate that all required fields are mapped"""
        if not mapping:
            return False

        required_fields = set(DataProcessor.REQUIRED_FIELDS)
        provided_keys = set(mapping.keys())
        if not required_fields.issubset(provided_keys):
            return False

        for field in DataProcessor.REQUIRED_FIELDS:
            value = mapping.get(field)
            if not value:
                return False

        return True

    @staticmethod
    def process_data(upload, mapping):
        """
        Process and validate uploaded data based on column mapping.
        Returns: (rows_list, quality_report, error)
        """
        try:
            rows = DataProcessor.read_file(upload.file, upload.file_format)

            # Rename columns according to mapping
            reverse_mapping = {v: k for k, v in mapping.items()}
            mapped_rows = []
            for row in rows:
                new_row = {}
                for orig_key, value in row.items():
                    if orig_key in reverse_mapping:
                        new_row[reverse_mapping[orig_key]] = value
                    else:
                        new_row[orig_key] = value
                mapped_rows.append(new_row)

            # Remove fully empty rows
            mapped_rows = [r for r in mapped_rows if any(v is not None and v != '' for v in r.values())]

            quality_report = {
                'total_rows': len(mapped_rows),
                'missing_values': {},
                'data_types': {},
                'warnings': [],
                'invalid_dates': 0,
                'invalid_numeric': 0,
            }

            # Count missing values per field
            for field in DataProcessor.REQUIRED_FIELDS:
                missing_count = sum(1 for r in mapped_rows if r.get(field) is None or r.get(field) == '')
                quality_report['missing_values'][field] = missing_count

            # Infer data types from first non-null value
            for field in DataProcessor.REQUIRED_FIELDS:
                for r in mapped_rows:
                    val = r.get(field)
                    if val is not None and val != '':
                        if _coerce_numeric(val) is not None:
                            quality_report['data_types'][field] = 'float64'
                        else:
                            quality_report['data_types'][field] = 'object'
                        break
                else:
                    quality_report['data_types'][field] = 'unknown'

            # Validate numeric columns
            numeric_cols = ['BS&W (%)', 'Net Oil (bopd)', 'Form.GLR (scf/bbl)',
                          'Tubing Pressure (psi)', 'Flow Line Pressure (psi)']

            invalid_numeric_cols = []
            for col in numeric_cols:
                for r in mapped_rows:
                    val = r.get(col)
                    if val is not None and val != '' and _coerce_numeric(val) is None:
                        invalid_numeric_cols.append(col)
                        quality_report['invalid_numeric'] += 1
                        break

            if invalid_numeric_cols:
                quality_report['warnings'].append(
                    f"Could not convert some values in the following numeric columns: {', '.join(sorted(set(invalid_numeric_cols)))}"
                )

            # Validate dates
            invalid_dates = 0
            for r in mapped_rows:
                val = r.get('Date')
                if val is not None and val != '':
                    try:
                        _parse_date(val)
                    except (ValueError, TypeError):
                        invalid_dates += 1

            quality_report['invalid_dates'] = invalid_dates
            if invalid_dates > 0:
                quality_report['warnings'].append(
                    f"Could not parse {invalid_dates} date(s). Please use a supported date format."
                )

            # Flag missing required fields in sample data
            missing_required = [field for field, count in quality_report['missing_values'].items() if count > 0]
            if missing_required:
                quality_report['warnings'].append(
                    f"Required fields missing values in: {', '.join(missing_required)}"
                )

            if not mapped_rows:
                return None, None, 'No valid rows were found in the uploaded file after mapping.'

            return mapped_rows, quality_report, None

        except Exception as e:
            logger.error(f"Error processing data: {str(e)}")
            return None, None, str(e)

    @staticmethod
    def create_preview(upload, mapping):
        """Create preview data (first 50 rows)"""
        rows, quality_report, error = DataProcessor.process_data(upload, mapping)

        if error:
            return None, error

        # Sample first 50 rows, converting any non-serializable types
        sample = rows[:50]
        for row in sample:
            for key, value in row.items():
                if isinstance(value, datetime):
                    row[key] = value.isoformat()
                elif isinstance(value, (bytes, bytearray)):
                    row[key] = str(value)
                elif value is None:
                    row[key] = None

        # Create preview record
        preview = PreviewData.objects.create(
            upload=upload,
            sample_data=sample,
            data_quality_report=quality_report
        )

        return preview, None


def _parse_date(value):
    """Parse a date string into a datetime object. Raises ValueError on failure."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Could be Excel serial date number
        from datetime import timedelta
        excel_epoch = datetime(1899, 12, 30)
        return excel_epoch + timedelta(days=float(value))

    value = str(value).strip()
    # Try common date formats
    formats = [
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%m-%d-%Y',
        '%d-%m-%Y',
        '%Y%m%d',
        '%m/%d/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%d-%b-%Y',  # e.g., 01-Jan-2024
        '%d-%b-%y',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    raise ValueError(f"Unable to parse date: {value}")