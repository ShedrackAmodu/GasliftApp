from django.shortcuts import render, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from django.http import HttpResponse
from apps.analysis.models import AnalysisSession, WellTrendAnalysis, CompletionData
from apps.analysis.reporting import ReportGenerator

@login_required(login_url='accounts:login')
@require_http_methods(["GET"])
def view_results(request, analysis_id):
    """View analysis results - Step 7"""
    analysis = get_object_or_404(AnalysisSession, id=analysis_id, user=request.user)
    
    # Get paginated results
    well_trends = WellTrendAnalysis.objects.filter(analysis=analysis).order_by('rank')
    
    paginator = Paginator(well_trends, 25)  # 25 per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get completion data for context
    completion_data = {}
    for completion in CompletionData.objects.filter(analysis=analysis):
        completion_data[completion.well_id] = completion
    
    # Annotate well trends with completion feasibility
    for trend in page_obj.object_list:
        trend.completion_info = completion_data.get(trend.well_id)
    
    context = {
        'analysis': analysis,
        'page_obj': page_obj,
        'well_trends': page_obj.object_list,
        'total_wells': well_trends.count(),
        'completion_data': completion_data,
    }
    
    return render(request, 'results/view_results.html', context)

@login_required(login_url='accounts:login')
@require_http_methods(["GET"])
def export_excel(request, analysis_id):
    """Export results to Excel - Step 8"""
    analysis = get_object_or_404(AnalysisSession, id=analysis_id, user=request.user)
    
    well_trends = WellTrendAnalysis.objects.filter(analysis=analysis).order_by('rank')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gas Lift Candidates"
    
    # Headers matching the manual's specification
    headers = [
        'Rank',
        'Well',
        'UID',
        'BSW_Flag',
        'OilRate_Flag',
        'GLR_Flag',
        'BSW Trend',
        'Oil Rate Trend',
        'GLR Trend',
        'Candidate Score',
        'Prod Method',
        'Test Status',
        'Flow Line Pressure (psi)',
        'Well Choke Size',
        'Data Quality Score',
        'Outliers Removed',
        'Choke Normalized',
        'Liquid Loading Flag',
        'Critical Velocity (ft/s)',
        'Days to Economic Limit',
        'Recommended Gas (MMscf/d)',
        'Completion Feasibility',
        'Summary Comment',
    ]
    
    ws.append(headers)
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for trend in well_trends:
        ws.append([
            trend.rank,
            trend.well_id,
            str(trend.id),
            'Yes' if trend.bsw_flag else 'No',
            'Yes' if trend.oil_rate_flag else 'No',
            'Yes' if trend.glr_flag else 'No',
            trend.bsw_trend or '',
            trend.oil_rate_trend or '',
            trend.glr_trend or '',
            round(trend.candidate_score, 2),
            trend.prod_method or '',
            trend.test_status or '',
            trend.flow_line_pressure if trend.flow_line_pressure is not None else '',
            trend.well_choke_size or '',
            trend.data_quality_score,
            trend.outlier_count,
            'Yes' if trend.is_choke_normalized else 'No',
            'Yes' if trend.liquid_loading_flag else 'No',
            trend.critical_velocity if trend.critical_velocity is not None else '',
            trend.days_to_economic_limit if trend.days_to_economic_limit is not None else '',
            trend.recommended_gas_mmscf if trend.recommended_gas_mmscf is not None else '',
            trend.get_completion_feasibility_display() if trend.completion_feasibility else 'Unknown',
            trend.summary_comment,
        ])
    
    # Adjust column widths
    column_widths = {
        'A': 8, 'B': 15, 'C': 38, 'D': 10, 'E': 13,
        'F': 10, 'G': 12, 'H': 15, 'I': 12, 'J': 15,
        'K': 15, 'L': 15, 'M': 20, 'N': 15, 'O': 18,
        'P': 16, 'Q': 16, 'R': 16, 'S': 18, 'T': 22,
        'U': 22, 'V': 25, 'W': 40,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="GasLift_Candidates.xlsx"'
    
    wb.save(response)
    return response

@login_required(login_url='accounts:login')
@require_http_methods(["GET"])
def export_csv(request, analysis_id):
    """Export results to CSV"""
    analysis = get_object_or_404(AnalysisSession, id=analysis_id, user=request.user)
    
    well_trends = WellTrendAnalysis.objects.filter(analysis=analysis).order_by('rank')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="GasLift_Candidates.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Rank', 'Well', 'UID', 'BSW_Flag', 'OilRate_Flag', 'GLR_Flag',
        'BSW Trend', 'Oil Rate Trend', 'GLR Trend', 'Candidate Score',
        'Prod Method', 'Test Status', 'Flow Line Pressure (psi)',
        'Well Choke Size', 'Data Quality Score', 'Outliers Removed',
        'Choke Normalized', 'Liquid Loading Flag', 'Critical Velocity (ft/s)',
        'Days to Economic Limit', 'Recommended Gas (MMscf/d)',
        'Completion Feasibility', 'Summary Comment',
    ])
    
    for trend in well_trends:
        writer.writerow([
            trend.rank,
            trend.well_id,
            str(trend.id),
            'Yes' if trend.bsw_flag else 'No',
            'Yes' if trend.oil_rate_flag else 'No',
            'Yes' if trend.glr_flag else 'No',
            trend.bsw_trend or '',
            trend.oil_rate_trend or '',
            trend.glr_trend or '',
            round(trend.candidate_score, 2),
            trend.prod_method or '',
            trend.test_status or '',
            trend.flow_line_pressure if trend.flow_line_pressure is not None else '',
            trend.well_choke_size or '',
            trend.data_quality_score,
            trend.outlier_count,
            'Yes' if trend.is_choke_normalized else 'No',
            'Yes' if trend.liquid_loading_flag else 'No',
            trend.critical_velocity if trend.critical_velocity is not None else '',
            trend.days_to_economic_limit if trend.days_to_economic_limit is not None else '',
            trend.recommended_gas_mmscf if trend.recommended_gas_mmscf is not None else '',
            trend.get_completion_feasibility_display() if trend.completion_feasibility else 'Unknown',
            trend.summary_comment,
        ])
    
    return response

@login_required(login_url='accounts:login')
@require_http_methods(["GET"])
def filter_results(request, analysis_id):
    """API endpoint for filtering results"""
    analysis = get_object_or_404(AnalysisSession, id=analysis_id, user=request.user)
    
    well_trends = WellTrendAnalysis.objects.filter(analysis=analysis)
    
    # Apply filters
    bsw_flag = request.GET.get('bsw_flag')
    oil_flag = request.GET.get('oil_flag')
    glr_flag = request.GET.get('glr_flag')
    search = request.GET.get('search', '')
    prod_method = request.GET.get('prod_method', '')
    
    if bsw_flag:
        well_trends = well_trends.filter(bsw_flag=bsw_flag.lower() == 'true')
    if oil_flag:
        well_trends = well_trends.filter(oil_rate_flag=oil_flag.lower() == 'true')
    if glr_flag:
        well_trends = well_trends.filter(glr_flag=glr_flag.lower() == 'true')
    if search:
        well_trends = well_trends.filter(Q(well_id__icontains=search) | Q(summary_comment__icontains=search))
    if prod_method:
        well_trends = well_trends.filter(prod_method__icontains=prod_method)
    
    well_trends = well_trends.order_by('rank')
    
    data = {
        'results': [
            {
                'rank': wt.rank,
                'well_id': wt.well_id,
                'uid': str(wt.id),
                'bsw_flag': wt.bsw_flag,
                'oil_rate_flag': wt.oil_rate_flag,
                'glr_flag': wt.glr_flag,
                'candidate_score': wt.candidate_score,
                'prod_method': wt.prod_method or '',
                'test_status': wt.test_status or '',
                'flow_line_pressure': wt.flow_line_pressure,
                'well_choke_size': wt.well_choke_size or '',
                'summary_comment': wt.summary_comment,
                'bsw_trend': wt.bsw_trend or '',
                'oil_rate_trend': wt.oil_rate_trend or '',
                'glr_trend': wt.glr_trend or '',
            }
            for wt in well_trends
        ]
    }
    
    return HttpResponse(json.dumps(data), content_type='application/json')

@login_required(login_url='accounts:login')
@require_http_methods(["GET"])
def export_pdf(request, analysis_id):
    """Export results as executive summary PDF"""
    analysis = get_object_or_404(AnalysisSession, id=analysis_id, user=request.user)
    well_trends = WellTrendAnalysis.objects.filter(analysis=analysis).order_by('rank')
    
    # Get completion data
    completion_data = {}
    for completion in CompletionData.objects.filter(analysis=analysis):
        completion_data[completion.well_id] = completion
    
    pdf_bytes = ReportGenerator.generate_pdf(analysis, well_trends, completion_data)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="GasLift_Executive_Summary_{str(analysis_id)[:8]}.pdf"'
    response.write(pdf_bytes)
    return response
